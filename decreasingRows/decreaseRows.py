# - Main imports -
from openpyxl import Workbook, load_workbook
from geopy.distance import great_circle
from geopy.distance import geodesic
import sys

# LOGIC:
  # 0) Copy the current data sheet into a new sheet that will be modified
  # 1) Find a zip code with the lowest % in the table
  # 2) Delete the zipcode from Step 1 from the table but save its data
  # 3) Find the closest zip code to that zipcode 
    # 3.1) Go over each zipcode and find distance between them
    # 3.2) Keep track of the lowest distance
    # 3.3) Return the zipcode, weight, row number of the lowest distance
  # 4) Add the weight of the zip from Step 1 to the closest zip code from Step 2. Edit hisotory of merging.
  # 5) Do 1-4 until certain number of rows is reached.


# ! Settings:
teuWBname = 'data13.xlsx' # Name of Excel file that holds TEUs
teuTableName = 'Teu' # Name of Table that holds TEU with zipcodes
columnZipTeu = 'A' # Column in teuTableName that holds zipcodes
columnTeu = 'C' # Column in teuTableName that holds TEU weights
numberOfUselessTopRows = 1 # Number of useless rows on top of TEU table
numberOfUselessBottomRows = 2 # Number of useless rows on bottom of TEU table

coordinatesWBname = 'data13.xlsx' # Name of Excel file that holds long and lat
coordinatesTableName = 'zip_code_database' # Name of Table that holds long and lat with zipcodes
columnLatCoord = 'M' # Column in coordinatesTableName that holds Latitudes
columnLongCoord = 'N' # Column in coordinatesTableName that holds Longitudes

startOver = False # True = combine rows from the main TEU table; False = combine rows from the existing "mergedZipCodes" table




# -------------------------------------------------------------------------------------------------------------------- #


# - Function for finding a distance between two coordinates -
def getDistance(location1, location2):
  return geodesic(location1, location2).miles

# - Function for finding coordinates of a given zipcode in ZipByKyle sheet -
def findCoordinates(zipcode, wb):
  coordinatesDataSheet = wb[coordinatesTableName]
  
  for row in coordinatesDataSheet.iter_rows(min_row=None, max_row=None, min_col=1, max_col=1, values_only=False):
    for cell in row:
      if( isinstance(cell.value, int) ):
        if( cell.value == zipcode ):
          return (coordinatesDataSheet[columnLatCoord + str(cell.row)].value, coordinatesDataSheet[columnLongCoord + str(cell.row)].value)
  # if nothing found
  return (-1, -1)


# -------------------------------------------------------------------------------------------------------------------- #

# - Load the Excel file -
wb = load_workbook(teuWBname)
coordinatesSrc = load_workbook(coordinatesWBname, read_only=True)

# - Make sure mergedZipCodes sheet doesn't exist -
if( 'mergedZipCodes' in wb.sheetnames and startOver ):
  del wb['mergedZipCodes']

if( startOver ):
  # - Copy the Top457-Dist sheet -
  sortedData = wb.copy_worksheet(wb[teuTableName])
  sortedData.title = 'mergedZipCodes'
else :
  sortedData = wb['mergedZipCodes']

# - Find range -
decreaseUntil = int(input("Please, enter the number to which you'd like to decrease number of rows in the table: "))
decreaseUntil = decreaseUntil + numberOfUselessTopRows + numberOfUselessBottomRows

# - Iterate -

# Find the last row with useful data
currentIterator = sortedData.max_row - numberOfUselessBottomRows

while( sortedData.max_row > decreaseUntil ):
  # Find and save values of the row
  currentZipCode = sortedData[columnZipTeu + str(currentIterator)].value
  currentCoordinates = findCoordinates(currentZipCode, coordinatesSrc)
  if(currentCoordinates == (-1,-1)):
    sortedData.delete_rows(currentIterator)
    currentIterator -= 1
    
    print(currentZipCode, " can't be found in coordinates source excel file: coordinatesWBname [outer while loop], currentIterator = ", currentIterator)
    continue

  currentHistory = ''
  if( not sortedData['RY' + str(currentIterator)].value or sortedData['RY' + str(currentIterator)].value == None):
    currentHistory = str(currentZipCode)
  else:
    currentHistory = '[(' + str(sortedData['RY' + str(currentIterator)].value) + ')' + '->' + str(currentZipCode) + ']'

  currentWeight = sortedData[columnTeu + str(currentIterator)].value

  # Delete the row, shift currentIterator to the upper row
  sortedData.delete_rows(currentIterator)
  currentIterator -= 1

  # Prepare variables for finding a zipcode with the lowest distance
  lowestDistance = sys.maxsize
  rowOfLD = -1

  # Find zipcode with the lowest distance
  for i in range(1, sortedData.max_row+1):
      zipCode = sortedData[columnZipTeu + str(i)].value

      if(isinstance(zipCode, int)):
        holder = findCoordinates(zipCode, coordinatesSrc)
        if(holder == (-1,-1)):
          print(zipCode, " can't be found in coordinates source excel file [inner for loop]")
          continue
        
        holder = getDistance(currentCoordinates, findCoordinates(zipCode, coordinatesSrc))

        if( holder < lowestDistance):
          lowestDistance = holder
          rowOfLD = i

  # Add weight of the deleted row to the found row
  sortedData[columnTeu + str(rowOfLD)].value = sortedData[columnTeu + str(rowOfLD)].value + currentWeight

  # Acommodate history of the changed row
  historyColLD = sortedData['RY' + str(rowOfLD)].value
  if ( not historyColLD or historyColLD == None ):
    sortedData['RY' + str(rowOfLD)].value = currentHistory
  else:
    sortedData['RY' + str(rowOfLD)].value = str(historyColLD) + ' -> ' + currentHistory

  print(rowOfLD) # to be deleted


# - Save changes -
wb.save('data.xlsx')

