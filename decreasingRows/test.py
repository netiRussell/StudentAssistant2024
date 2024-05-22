from openpyxl import Workbook, load_workbook

def findCoordinates(zipcode, wb):
  coordinatesDataSheet = wb['RawData']
  # for i in range(1, coordinatesDataSheet.max_row+1):
  #   if(isinstance(coordinatesDataSheet['A' + str(i)].value, int)):
  #     if( coordinatesDataSheet['A' + str(i)].value == zipcode):
  #       return (coordinatesDataSheet['D' + str(i)].value, coordinatesDataSheet['E' + str(i)].value)
  
  for row in coordinatesDataSheet.iter_rows(min_row=None, max_row=None, min_col=1, max_col=1, values_only=False):
    for cell in row:
      if( isinstance(cell.value, int) ):
        if( cell.value == zipcode ):
          return (coordinatesDataSheet['D' + str(cell.row)].value, coordinatesDataSheet['E' + str(cell.row)].value)
  # if nothing found
  return (-1, -1)


coordinatesSrc = load_workbook('CAcoordinates.xlsx', read_only=True)

if( findCoordinates(91607, coordinatesSrc) == (-1,-1) ):
  print("can't find this zip")