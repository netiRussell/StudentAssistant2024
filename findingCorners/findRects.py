import cv2
import ast
import easyocr
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

# - Create an Excel file that will contain all the coordinates -
wb = Workbook()
ws = wb.active
ws.append(['Top left X', 'Top left Y', 'Width', 'Height'])

# - Get img and prepare it for contour detection -
img = cv2.imread('img/sample.png')
heightImg, widthImg, channels = img.shape
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
ret,thresh = cv2.threshold(gray,50,255,0)

# - Find contours -
contours,hierarchy = cv2.findContours(thresh, 1, cv2.CHAIN_APPROX_SIMPLE )

# - Gather data -
for cnt in contours:
   x1,y1 = cnt[0][0]
   approx = cv2.approxPolyDP(cnt, 0.01*cv2.arcLength(cnt, True), True)

   # If found a rectangle
   if len(approx) == 4:
      # and that rectangle is not too small
      if(cv2.contourArea(approx) < 10):
         continue

      x, y, w, h = cv2.boundingRect(cnt)

      # Disregard a rectangle that matches the screen (bad implementation perfomance wise)
      if( w == widthImg and h == heightImg):
         continue

      ws.append([x, y, w, h])


# - Clean up data -
# xy1 and xy2 are dupicates of each other if the difference of w1-w2 == h1-h2
for mainRow in ws.iter_rows(min_row=2, max_row=None, min_col=1, max_col=4, values_only=False):
   currentX = mainRow[0].value
   currentY = mainRow[1].value
   currentH = mainRow[2].value
   currentW = mainRow[3].value
   for checkingRow in ws.iter_rows(min_row=mainRow[0].row+1, max_row=None, min_col=1, max_col=4, values_only=False):
      if(checkingRow[0].value != None ):
         if(currentH - checkingRow[2].value == currentW - checkingRow[3].value):
            if(currentX - checkingRow[0].value < 10 and currentY - checkingRow[1].value < 10):
               ws.delete_rows(mainRow[0].row)
               break

# - Create coordinates -
coordinatesWS = wb.create_sheet()
coordinatesWS.title = 'Coordinates'
coordinatesWS.append(['Top left (X,Y)', 'Top right', 'Bottom left', 'Bottom right', 'Department'])
reader = easyocr.Reader(['en'], gpu=False)
gray = cv2.blur(gray, (2,2))

for row in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=False):
   if(row[0].value == None):
      continue

   # Detect department number
   x,y,w,h = row[0].value, row[1].value, row[2].value, row[3].value
   ROI = gray[y:y+h,x:x+w]
   text_ = reader.readtext(ROI, mag_ratio=2, allowlist="0123456789", text_threshold=0.5)

   deptNumber = ''
   for t_, t in enumerate(text_):
      bbox, deptNumber, score = t
      print(deptNumber)

      if score > 0.25:
         break
   
   # Insert values to the row
   coordinatesWS['A' + str(row[0].row)] = '(' + str(row[0].value) + ', ' + str(row[1].value) + ')'
   coordinatesWS['B' + str(row[0].row)] = '(' + str(row[0].value + row[2].value) + ', ' + str(row[1].value) + ')'
   coordinatesWS['C' + str(row[0].row)] = '(' + str(row[0].value) + ', ' + str(row[1].value + row[3].value) + ')'
   coordinatesWS['D' + str(row[0].row)] = '(' + str(row[0].value + row[2].value) + ', ' + str(row[1].value + row[3].value) + ')'
   coordinatesWS['E' + str(row[0].row)] = int(deptNumber)


# - Save the Excel file -
wb.save('coordinates.xlsx')

# - Draw each point from the final data -
x_data = np.array([], dtype=np.float32)
y_data = np.array([], dtype=np.float32)
for row in coordinatesWS.iter_rows(min_row=2, max_row=None, min_col=None, max_col=4, values_only=False):
   for cell in row:
      holder = ast.literal_eval(cell.value)
      img[holder[1], holder[0]] = [0,0,255]
      
      x_data = np.append(x_data, holder[0])
      y_data = np.append(y_data, holder[1])

# - Plot the given points -
fig, (ax1, ax2) = plt.subplots(1, 2)
fig.suptitle('Plots')
ax1.scatter(x_data, y_data)
ax1.set_ylim(bottom=heightImg, top=0 )
ax1.set_xlim(left=0, right=widthImg )
ax2.imshow(img)
plt.show()

# - Show img with point -
# cv2.imshow("Shapes", img)
# cv2.waitKey(0)
# cv2.destroyAllWindows()